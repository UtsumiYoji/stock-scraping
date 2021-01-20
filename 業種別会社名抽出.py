#Yahooか抽出　https://stocks.finance.yahoo.co.jp/stocks/qi/?ids=0050

import web_scraping
import openpyxl

def main():
    print("1ページ目のURLを入力して下さい")
    html = input()
    target_html = html

    i = 1
    while True:
        #解析対象を指定
        print(str(i) + 'ページ目のURL:' + target_html)
        text = web_scraping.get_text(target_html)
        text = text.split("\n")

        #ページに続きがないようなら終わり
        if len(text) == 67:
            print("終わります")
            break

        #業種を定義
        industry = text[0].replace("業種別銘柄一覧：", "")
        industry = industry.replace(" - Yahoo!ファイナンス", "")

        #Excelを開く
        wb = openpyxl.open('./業種別一覧.xlsx')

        #業種名のシートがないなら追加する
        if not industry in wb.sheetnames:
            wb.create_sheet(title=industry)
            sh = wb[industry]

            #見出しの作成
            sh['A1'].value = "コード"
            sh['B1'].value = "市場"
            sh['C1'].value = "社名"
        
        #シートを定義
        sh = wb[industry]
        
        #最大20回を40の位置から3飛ばしで繰り返す
        for j in range(40, 100, 3):
            #会社名がないなら終わる
            if not "株" in text[j]:
                break

            #会社名があるならExcelシートに加える
            write_row = sh.max_row + 1
            sh.cell(row=write_row, column=1).value = int(text[j][:4])

            #東証JQSだけ文字数違うから特別処理
            if "東証JQS" in text[j]:
                sh.cell(row=write_row, column=2).value = "東証JQS"
                sh.cell(row=write_row, column=3).value = text[j][9:]

            else:
                sh.cell(row=write_row, column=2).value = text[j][4:8]
                sh.cell(row=write_row, column=3).value = text[j][8:]

        #Excelの保存
        wb.save('./業種別一覧.xlsx')

        #次のページのURLを定義する
        i += 1
        target_html = html + '&p=' + str(i)

if __name__ == "__main__":
    while True:
        main()