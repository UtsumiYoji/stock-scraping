import web_scraping
import openpyxl

def main():
    print("1ページ目のURLを入力して下さい")
    html = input()
    target_html = html

    i = 0
    while True:
        #ページを更新
        i += 1
        target_html = html + '&p=' + str(i)

        #解析対象を指定
        print(str(i) + 'ページ目のURL:' + target_html)
        text = web_scraping.get_slash_text(target_html)
        text = text.split("\n")
        
        for x in range(len(text)):
            print(str(x) + ' ' + text[x])

        #ページに続きがないようなら終わり
        if text[228] == '/':
            print("終わります")
            break

        #Excelを開く
        wb = openpyxl.open('./業種別一覧.xlsx')
        sh_num = len(wb.sheetnames)

        #会社ごとにリスト化
        roa_data = text[228].replace('(連)', '(単)')
        roa_data = roa_data.split('掲示板')

        #最後まで繰り返す
        for j in range(len(roa_data)-1):
            target_text = roa_data[j].split('/')

            #コードとroaを覚える
            code = int(target_text[2])
            roa = target_text[7]

            #一致するコードを探す
            for k in range(sh_num):
                sh = wb.worksheets[k]
                for y in range(1, sh.max_row+1):
                    if sh.cell(row=y, column=1).value == code:
                        sh.cell(row=y, column=5).value = roa
                        break
        
        #保存
        wb.save('./業種別一覧.xlsx')

if __name__ == "__main__":
    main()